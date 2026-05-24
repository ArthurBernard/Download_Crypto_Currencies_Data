#!/usr/bin/env python3
# coding: utf-8

import sqlite3

import polars as pl
import pytest

from dccd.tools.io import IODataBase, get_df, save_df

_DF = pl.DataFrame({'a': [1, 2], 'b': [3.0, 4.0]})


def test_invalid_method(tmp_data_path):
    with pytest.raises(NotImplementedError):
        IODataBase(tmp_data_path, 'unknown')


# --- DataFrame (pickle) ---

def test_save_as_dataframe_create(tmp_data_path):
    db = IODataBase(tmp_data_path, 'dataframe')
    db.save_as_dataframe(_DF, name='test')
    result = db.get_from_dataframe('test')
    assert list(result.columns) == ['a', 'b']
    assert len(result) == 2


def test_save_as_dataframe_append(tmp_data_path):
    db = IODataBase(tmp_data_path, 'dataframe')
    db.save_as_dataframe(_DF, name='test')
    db.save_as_dataframe(_DF, name='test')
    result = db.get_from_dataframe('test')
    assert len(result) == 4


def test_get_df_missing(tmp_data_path):
    result = get_df(tmp_data_path, 'nonexistent')
    assert isinstance(result, pl.DataFrame)
    assert len(result) == 0


def test_save_load_df(tmp_data_path):
    save_df(_DF, tmp_data_path, 'roundtrip')
    result = get_df(tmp_data_path, 'roundtrip')
    assert list(result.columns) == list(_DF.columns)
    assert len(result) == len(_DF)


# --- CSV ---

def test_save_as_csv_create(tmp_data_path):
    db = IODataBase(tmp_data_path, 'csv')
    db.save_as_csv(_DF, name='test')
    import os
    path = tmp_data_path + '/test.csv'
    assert os.path.exists(path)
    loaded = pl.read_csv(path)
    assert list(loaded.columns) == ['a', 'b']
    assert len(loaded) == 2


def test_save_as_csv_append_no_duplicate_header(tmp_data_path):
    db = IODataBase(tmp_data_path, 'csv')
    db.save_as_csv(_DF, name='test')
    db.save_as_csv(_DF, name='test')
    loaded = pl.read_csv(tmp_data_path + '/test.csv')
    assert len(loaded) == 4


# --- SQLite ---

def test_save_as_sqlite(tmp_data_path):
    db = IODataBase(tmp_data_path, 'sqlite')
    db.save_as_sqlite(_DF, name='test', table='data')
    conn = sqlite3.connect(tmp_data_path + '/test.db')
    result = pl.read_database('SELECT * FROM data', connection=conn)
    conn.close()
    assert 'a' in result.columns
    assert len(result) == 2


# --- Excel ---

def test_save_as_excel(tmp_data_path):
    import openpyxl
    db = IODataBase(tmp_data_path, 'excel')
    db.save_as_excel(_DF, name='test', sheet_name='Sheet1')
    wb = openpyxl.load_workbook(tmp_data_path + '/test.xlsx')
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ('a', 'b')   # header
    assert len(rows) == 3          # 1 header + 2 data rows


def test_save_as_excel_append(tmp_data_path):
    import openpyxl
    db = IODataBase(tmp_data_path, 'excel')
    db.save_as_excel(_DF, name='test', sheet_name='Sheet1')
    db.save_as_excel(_DF, name='test', sheet_name='Sheet1')
    wb = openpyxl.load_workbook(tmp_data_path + '/test.xlsx')
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 5          # 1 header + 4 data rows


# --- __call__ dispatch ---

def test_call_dispatches_csv(tmp_data_path):
    db = IODataBase(tmp_data_path, 'csv')
    db(_DF, name='dispatch_test')
    import os
    assert os.path.exists(tmp_data_path + '/dispatch_test.csv')


# --- Parquet ---

def test_save_as_parquet_create(tmp_data_path):
    db = IODataBase(tmp_data_path, 'parquet')
    db.save_as_parquet(_DF, name='test')
    result = pl.read_parquet(tmp_data_path + '/test.parquet')
    assert list(result.columns) == ['a', 'b']
    assert len(result) == 2


def test_save_as_parquet_append(tmp_data_path):
    db = IODataBase(tmp_data_path, 'parquet')
    db.save_as_parquet(_DF, name='test')
    db.save_as_parquet(_DF, name='test')
    result = pl.read_parquet(tmp_data_path + '/test.parquet')
    assert len(result) == 4


# --- 'polars' method alias ---

def test_polars_method_alias(tmp_data_path):
    db = IODataBase(tmp_data_path, 'polars')
    db(_DF, name='test')
    result = pl.read_parquet(tmp_data_path + '/test.parquet')
    assert 'a' in result.columns
    assert len(result) == 2
