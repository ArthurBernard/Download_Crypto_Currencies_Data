#!/usr/bin/env python3
# coding: utf-8
# @Author: ArthurBernard
# @Email: arthur.bernard.92@gmail.com
# @Date: 2019-07-26 11:54:55
# @Last modified by: ArthurBernard

""" Tools and object to load, append and save different kind of database. """

# Built-in packages
import os.path
import time
from collections.abc import Callable
from os import makedirs
from pickle import Pickler, Unpickler
from typing import Any, Literal

# Third-party packages
import polars as pl
from sqlalchemy import URL, create_engine

__all__ = ['IODataBase', 'get_df', 'save_df']


class IODataBase:
    """ Object to save a pl.DataFrame into different kind/format of database.

    Parameters
    ----------
    path : str, optional
        Path of the database, default is './' (current directory).
    method : {'DataFrame', 'SQLite', 'CSV', 'Excel', 'parquet', 'PostgreSQL',\
              'Oracle', 'MSSQL', 'MySQL'}
        Format of database, default is CSV.

    Attributes
    ----------
    path : str
        Path of the database.
    method : str
        Kind/format of the database.

    Methods
    -------
    save_as_dataframe
    save_as_sql
    save_as_sqlite
    save_as_csv
    save_as_excel
    save_as_parquet
    __call__

    """

    def __init__(self, path: str = './', method: str = 'csv') -> None:
        makedirs(path, exist_ok=True)
        self.path = path if path.endswith('/') else path + '/'
        self.method = method.lower()
        self.parser: dict[str, Callable[..., None]] = {
            'dataframe': self.save_as_dataframe,
            'sqlite': self.save_as_sqlite,
            'csv': self.save_as_csv,
            'excel': self.save_as_excel,
            'parquet': self.save_as_parquet,
            'polars': self.save_as_parquet,
            'postgresql': self.save_as_sql,
            'mysql': self.save_as_sql,
            'oracle': self.save_as_sql,
            'mssql': self.save_as_sql,
        }
        if self.method not in self.parser:
            raise NotImplementedError(
                f"`method` must be one of {list(self.parser)}, got {method!r}"
            )

    def __call__(self, new_data: pl.DataFrame, **kwargs: Any) -> None:
        """ Append and save *new_data* in the configured format. """
        return self.parser[self.method](new_data, **kwargs)

    def save_as_dataframe(self, new_data: pl.DataFrame, name: str | None = None, ext: str = '.dat') -> None:
        """ Append and save *new_data* as a pickle binary file. """
        if name is None:
            name = time.strftime('%y-%m-%d', time.gmtime(time.time()))
        existing = get_df(self.path, name, ext=ext)
        combined = pl.concat([existing, new_data]) if len(existing) > 0 else new_data
        save_df(combined, self.path, name, ext=ext)

    def get_from_dataframe(self, name: str, ext: str = '.dat') -> pl.DataFrame:
        """ Load data from pickle binary file. """
        return get_df(self.path, name, ext=ext)

    def save_as_sqlite(self, new_data: pl.DataFrame, table: str = 'main_table', name: str | None = None, ext: str = '.db') -> None:
        """ Append *new_data* into a SQLite table (stdlib sqlite3, no extra deps). """
        import sqlite3

        _PL_TO_SQLITE: dict[type, str] = {
            pl.Int8: 'INTEGER', pl.Int16: 'INTEGER', pl.Int32: 'INTEGER', pl.Int64: 'INTEGER',
            pl.UInt8: 'INTEGER', pl.UInt16: 'INTEGER', pl.UInt32: 'INTEGER', pl.UInt64: 'INTEGER',
            pl.Float32: 'REAL', pl.Float64: 'REAL',
            pl.Boolean: 'INTEGER',
        }

        if name is None:
            name = time.strftime('%y', time.gmtime(time.time()))
        path = self.path + name + ext
        conn = sqlite3.connect(path)
        cols = ', '.join(
            f'"{c}" {_PL_TO_SQLITE.get(type(t), "TEXT")}'
            for c, t in zip(new_data.columns, new_data.dtypes)
        )
        conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({cols})')
        placeholders = ', '.join(['?'] * len(new_data.columns))
        conn.executemany(f'INSERT INTO "{table}" VALUES ({placeholders})', new_data.iter_rows())
        conn.commit()
        conn.close()

    def get_from_sqlite(self, name: str, table: str = 'main_table', ext: str = '.db') -> pl.DataFrame:
        """ Load data from a SQLite table. """
        import sqlite3
        path = self.path + name + ext
        conn = sqlite3.connect(path)
        result = pl.read_database(f'SELECT * FROM "{table}"', connection=conn)
        conn.close()
        return result

    def save_as_sql(self, new_data: pl.DataFrame, table: str = 'main_table', name: str | None = None, ext: str = '', driver: str | None = None, username: str | None = None, password: str | None = None, host: str | None = None, port: str | int | None = None, **kwargs: Any) -> None:
        """ Append *new_data* into a SQL database via ADBC (PostgreSQL, MySQL, …).

        Requires the appropriate ADBC driver, e.g. ``adbc_driver_postgresql``.
        """
        if name is None:
            name = time.strftime('%y', time.gmtime(time.time()))
        drivername = self.method if driver is None else f"{self.method}+{driver}"
        url = URL.create(
            drivername, username=username, password=password, host=host,
            port=port, database=self.path + name + ext, query=kwargs,
        )
        engine = create_engine(url)
        new_data.write_database(table, connection=engine, if_table_exists='append', engine='adbc')

    def save_as_csv(self, new_data: pl.DataFrame, name: str | None = None, ext: str = '.csv') -> None:
        """ Append *new_data* to a CSV file (header written once). """
        if name is None:
            name = time.strftime('%y', time.gmtime(time.time()))
        path = self.path + name + ext
        if os.path.exists(path):
            with open(path, 'ab') as f:
                new_data.write_csv(f, include_header=False)
        else:
            new_data.write_csv(path)

    def save_as_parquet(self, new_data: pl.DataFrame, name: str | None = None, ext: str = '.parquet', compression: Literal['snappy', 'gzip', 'brotli', 'lz4', 'zstd'] = 'snappy') -> None:
        """ Append *new_data* to a Parquet file. """
        if name is None:
            name = time.strftime('%y', time.gmtime(time.time()))
        path = self.path + name + ext
        if os.path.exists(path):
            existing = pl.read_parquet(path)
            new_data = pl.concat([existing, new_data])
        new_data.write_parquet(path, compression=compression)

    def save_as_excel(self, new_data: pl.DataFrame, name: str | None = None, sheet_name: str = 'Sheet1', ext: str = '.xlsx') -> None:
        """ Append *new_data* to an Excel file (requires openpyxl).

        Warnings
        --------
        Slow method, not recommended for large datasets.

        """
        import openpyxl

        if name is None:
            name = time.strftime('%y-%m-%d', time.gmtime(time.time()))
        path = self.path + name + ext

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                write_header = False
            else:
                ws = wb.create_sheet(sheet_name)
                write_header = True
        else:
            wb = openpyxl.Workbook()
            ws = wb.active  # type: ignore[assignment]
            ws.title = sheet_name
            write_header = True

        if write_header:
            ws.append(new_data.columns)
        for row in new_data.iter_rows(named=False):
            ws.append(list(row))

        wb.save(path)


def get_df(path: str, name: str, ext: str = '') -> pl.DataFrame:
    """ Load a DataFrame from a pickle binary file.

    Parameters
    ----------
    path, name, ext : str
        Directory path, file name, and optional extension.

    Returns
    -------
    pl.DataFrame
        Loaded DataFrame, or an empty DataFrame if the file is not found.

    """
    if path[-1] != '/' and name[0] != '/':
        path += '/'
    if ext and not ext.startswith('.'):
        ext = '.' + ext
    try:
        with open(path + name + ext, 'rb') as f:
            return Unpickler(f).load()
    except FileNotFoundError:
        return pl.DataFrame()


def save_df(df: pl.DataFrame, path: str, name: str, ext: str = '') -> None:
    """ Save a DataFrame as a pickle binary file.

    Parameters
    ----------
    df : pl.DataFrame
        DataFrame to persist.
    path, name, ext : str
        Directory path, file name, and optional extension.

    """
    if path[-1] != '/' and name[0] != '/':
        path += '/'
    if ext and not ext.startswith('.'):
        ext = '.' + ext
    makedirs(path, exist_ok=True)
    with open(path + name + ext, 'wb') as f:
        Pickler(f).dump(df)
